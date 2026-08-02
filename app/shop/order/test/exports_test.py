import io

import pandas
import pytest
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rest_framework.fields import DateTimeField
from shop.order.exports import (
    COLUMN_WIDTH_PADDING,
    MAX_COLUMN_WIDTH,
    MIN_COLUMN_WIDTH,
    OrderExportSerializer,
    OrderProductExportSerializer,
    autofit_columns,
)
from shop.order.models import Order, OrderProductOptionRelation, OrderProductRelation
from shop.product.models import OptionGroup


@pytest.mark.django_db
def test_order_export_returns_dataframe_with_korean_renamed_columns(order_factory):
    completed_order = order_factory(status="completed")
    df = OrderExportSerializer(instance=Order.objects.filter(id=completed_order.id), many=True).export()
    assert df.to_dict(orient="records") == [
        {
            "주문 번호": str(completed_order.id),
            "주문 계정 이메일": completed_order.user.email,
            "고객명": "홍길동",
            "고객 전화번호": "01012345678",
            "고객 이메일": "customer@example.com",
            "고객 소속": None,
            "주문명": completed_order.name,
            "첫 결제 시간": DateTimeField().to_representation(completed_order.payment_histories.first().created_at),
            "첫 결제 금액": completed_order.first_paid_price,
            "현재 결제 금액": completed_order.current_paid_price,
            "현재 상태": "completed",
            "PortOne ID": "imp_test_completed",
        }
    ]


@pytest.mark.django_db
def test_order_export_returns_empty_dataframe_for_empty_queryset():
    # pandas 는 빈 data 에 대해 컬럼 없는 DataFrame 반환 — 행/열 모두 0.
    df = OrderExportSerializer(instance=Order.objects.none(), many=True).export()
    assert len(df) == 0
    assert df.empty


@pytest.mark.django_db
def test_order_product_export_flattens_options_as_dynamic_columns(ticket_product, order_factory):
    completed_order = order_factory(status="completed")
    size_group = OptionGroup.objects.create(product=ticket_product, name="사이즈")
    opr = completed_order.products.first()
    OrderProductOptionRelation.objects.create(
        order_product_relation=opr,
        product_option_group=size_group,
        product_option=size_group.options.create(name="M"),
    )
    OrderProductOptionRelation.objects.create(
        order_product_relation=opr,
        product_option_group=OptionGroup.objects.create(
            product=ticket_product,
            name="요청사항",
            is_custom_response=True,
            custom_response_pattern=r"^.*$",
        ),
        custom_response="배송 빠르게",
    )

    df = OrderProductExportSerializer(instance=OrderProductRelation.objects.filter(id=opr.id), many=True).export()
    # order_id / product_id 는 raw FK 라 UUID 그대로 노출 (DRF UUIDField 거치지 않음).
    assert df.to_dict(orient="records") == [
        {
            "주문 번호": completed_order.id,
            "상품 ID": ticket_product.id,
            "상품명": ticket_product.name,
            "상태": opr.status,
            "결제 금액": opr.price,
            "추가 기부액": opr.donation_price,
            "사이즈": "M",
            "요청사항": "배송 빠르게",
        }
    ]


@pytest.mark.django_db
def test_order_product_export_splits_repeated_options_of_one_group_into_numbered_columns(ticket_product, order_factory):
    # 한 주문 상품이 같은 그룹 옵션을 여러 개 고른 경우 — 한 컬럼에 덮어쓰면 두 번째부터 사라져
    # 사이즈별 수량 집계가 어긋난다 (재고 음수 사고의 원인).
    completed_order = order_factory(status="completed")
    size_group = OptionGroup.objects.create(product=ticket_product, name="검은색 티셔츠")
    small = size_group.options.create(name="S", priority=0)
    large = size_group.options.create(name="L", priority=10)
    opr = completed_order.products.first()
    for selected in (large, small, large):
        OrderProductOptionRelation.objects.create(
            order_product_relation=opr, product_option_group=size_group, product_option=selected
        )

    df = OrderProductExportSerializer(instance=OrderProductRelation.objects.filter(id=opr.id), many=True).export()
    row = df.to_dict(orient="records")[0]
    # priority 순 정렬이라 S 가 먼저, 그 다음 L 두 벌.
    assert row["검은색 티셔츠"] == "S"
    assert row["검은색 티셔츠 (2)"] == "L"
    assert row["검은색 티셔츠 (3)"] == "L"


@pytest.mark.django_db
def test_order_product_export_orders_option_columns_by_group_name(ticket_product, order_factory):
    # 옵션 컬럼은 행마다 동적으로 붙어 기본값이 "먼저 등장한 순" — 데이터 순서에 따라 `흰색 (2)` 가
    # `검은색 (3)` 앞으로 끼어들어 같은 그룹이 흩어진다. 고정 컬럼 뒤에 (그룹명, n) 순으로 배치한다.
    white = OptionGroup.objects.create(product=ticket_product, name="흰색 티셔츠")
    black = OptionGroup.objects.create(product=ticket_product, name="검은색 티셔츠")
    for group, counts in ((white, 2), (black, 3)):
        opr = order_factory(status="completed").products.get()
        selected = group.options.create(name="M")
        for _ in range(counts):
            OrderProductOptionRelation.objects.create(
                order_product_relation=opr, product_option_group=group, product_option=selected
            )

    df = OrderProductExportSerializer(instance=OrderProductRelation.objects.filter_active(), many=True).export()
    fixed = [label for _, label in OrderProductExportSerializer.Meta.field_def]
    assert list(df.columns) == [
        *fixed,
        "검은색 티셔츠",
        "검은색 티셔츠 (2)",
        "검은색 티셔츠 (3)",
        "흰색 티셔츠",
        "흰색 티셔츠 (2)",
    ]


@pytest.mark.django_db
def test_order_product_export_renders_unselected_optional_option_as_none(ticket_product, order_factory):
    # placeholder_mode=OPTIONAL 그룹은 product_option 없이 저장될 수 있다 — export 가 터지면 안 된다.
    completed_order = order_factory(status="completed")
    opr = completed_order.products.first()
    OrderProductOptionRelation.objects.create(
        order_product_relation=opr,
        product_option_group=OptionGroup.objects.create(
            product=ticket_product, name="사이즈", placeholder_mode=OptionGroup.PlaceholderMode.OPTIONAL
        ),
        product_option=None,
    )

    df = OrderProductExportSerializer(instance=OrderProductRelation.objects.filter(id=opr.id), many=True).export()
    assert df.to_dict(orient="records")[0]["사이즈"] is None


@pytest.mark.django_db
def test_order_product_export_calling_export_on_child_raises_to_force_list_serializer():
    # 단건 OrderProductExportSerializer.export() 는 NotImplemented — `many=True` 강제하는 guard.
    with pytest.raises(NotImplementedError):
        OrderProductExportSerializer().export()


@pytest.mark.django_db
def test_order_export_calling_export_on_child_raises():
    with pytest.raises(NotImplementedError):
        OrderExportSerializer().export()


def test_autofit_columns_sizes_each_column_to_its_widest_cell():
    long_header = "긴한글헤더입니다"
    df = pandas.DataFrame({"짧음": ["a"], long_header: ["x"], "값이긴열": ["가" * 40]})
    fileio = io.BytesIO()
    with pandas.ExcelWriter(fileio, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="s")
        autofit_columns(writer.sheets["s"], df)

    # xlsxwriter 는 같은 너비의 인접 열을 하나의 range 로 묶어 쓰므로 min~max 로 펼쳐서 읽는다.
    # 되읽은 너비에는 폰트 보정분(~0.71)이 붙어 abs=1 로 비교.
    widths = {
        get_column_letter(index): dim.width
        for dim in load_workbook(fileio)["s"].column_dimensions.values()
        for index in range(dim.min, dim.max + 1)
    }
    assert widths["A"] == pytest.approx(MIN_COLUMN_WIDTH, abs=1)  # index 열
    assert widths["B"] == pytest.approx(MIN_COLUMN_WIDTH, abs=1)  # 헤더 "짧음" 폭 4 → 하한
    # 헤더가 값보다 넓으면 헤더 기준. 한글은 2칸으로 센다.
    assert widths["C"] == pytest.approx(len(long_header) * 2 + COLUMN_WIDTH_PADDING, abs=1)
    assert widths["D"] == pytest.approx(MAX_COLUMN_WIDTH, abs=1)  # 값 폭 80 → 상한
